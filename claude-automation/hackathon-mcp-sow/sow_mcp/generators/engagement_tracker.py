import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..models import ScopeSection, SowDocument
from ..templates.excel_styles import (
    COLUMN_WIDTHS,
    HEADER_FILL,
    HEADER_FONT,
    SECTION_HEADER_FILL,
    SECTION_HEADER_FONT,
    THIN_BORDER,
    apply_data_row,
    apply_header,
)

_TRACKER_HEADERS = ["ID", "Title", "Type", "Owner", "Est. Hours", "Target Date", "Status", "Notes"]
_ALL_HEADERS = ["Section", "Section Title"] + _TRACKER_HEADERS


def _write_summary_sheet(wb: Workbook, doc: SowDocument) -> None:
    ws = wb.active
    ws.title = "Summary"

    rows: list[tuple[str, str]] = [
        ("Customer", doc.customer),
        ("Project Title", doc.project_title),
        ("Prepared By", doc.prepared_by),
        ("", ""),
        ("Delivery Phases", " → ".join(doc.phases) if doc.phases else "TBD"),
        ("Total VMs to Migrate", str(doc.vm_count) if doc.vm_count else "TBD"),
        ("Total Hypervisors", str(doc.hypervisor_count) if doc.hypervisor_count else "TBD"),
        ("Migration Tool", doc.migration_tool or "TBD"),
        ("", ""),
        ("Datacenters", ""),
    ]
    for dc in doc.datacenters:
        hv = f"{dc.hypervisor_count} hypervisors" if dc.hypervisor_count else "count TBD"
        rows.append(("", f"{dc.name}  ({dc.role}) — {hv}"))

    rows += [("", ""), ("Integrations", "")]
    for integration in doc.integrations:
        rows.append(("", integration))

    ws.column_dimensions["A"].width = COLUMN_WIDTHS["Label"]
    ws.column_dimensions["B"].width = COLUMN_WIDTHS["Value"]

    for row_idx, (label, value) in enumerate(rows, 1):
        lc = ws.cell(row=row_idx, column=1, value=label)
        vc = ws.cell(row=row_idx, column=2, value=value)
        if label:
            lc.font = HEADER_FONT
            lc.fill = HEADER_FILL
            lc.border = THIN_BORDER
            vc.border = THIN_BORDER
            vc.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = 18


def _write_section_sheet(wb: Workbook, section: ScopeSection) -> None:
    # Excel sheet names are max 31 chars
    sheet_title = f"{section.id} {section.title}"[:31]
    ws = wb.create_sheet(title=sheet_title)
    apply_header(ws, _TRACKER_HEADERS)

    if section.deliverables:
        for idx, d in enumerate(section.deliverables, 1):
            apply_data_row(
                ws,
                row_idx=idx + 1,
                values=[
                    d.id,
                    d.title,
                    d.type,
                    d.owner,
                    d.estimated_hours,
                    "",           # Target Date — to be filled by CE
                    "Not Started",
                    "",           # Notes
                ],
                is_odd=(idx % 2 == 1),
            )
    else:
        # No deliverables extracted — write one placeholder row so the sheet isn't empty
        ws.cell(row=2, column=1, value=f"D-{section.id}-1")
        ws.cell(row=2, column=2, value=f"[REVIEW] No deliverables extracted for this section")
        ws.cell(row=2, column=4, value="Platform9")
        ws.cell(row=2, column=7, value="Not Started")

    # Activities panel below the deliverables table
    if section.activities:
        gap = len(section.deliverables) + 3
        ws.cell(row=gap, column=1, value="Activities").font = SECTION_HEADER_FONT
        ws.cell(row=gap, column=1).fill = SECTION_HEADER_FILL
        for i, activity in enumerate(section.activities, 1):
            ws.cell(row=gap + i, column=1, value=f"{i}.").alignment = Alignment(horizontal="right")
            ac = ws.cell(row=gap + i, column=2, value=activity)
            ac.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[gap + i].height = max(15, len(activity) // 4)

    ws.freeze_panes = "A2"


def _write_all_deliverables_sheet(wb: Workbook, doc: SowDocument) -> None:
    ws = wb.create_sheet(title="All Deliverables")
    apply_header(ws, _ALL_HEADERS)

    # Widen the extra columns not covered by _TRACKER_HEADERS defaults
    ws.column_dimensions["A"].width = COLUMN_WIDTHS["Section"]
    ws.column_dimensions["B"].width = COLUMN_WIDTHS["Section Title"]

    row_idx = 2
    for section in doc.scope_sections:
        for d in section.deliverables:
            apply_data_row(
                ws,
                row_idx=row_idx,
                values=[
                    section.id,
                    section.title,
                    d.id,
                    d.title,
                    d.type,
                    d.owner,
                    d.estimated_hours,
                    "",           # Target Date
                    "Not Started",
                    "",           # Notes
                ],
                is_odd=(row_idx % 2 == 0),
            )
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_ALL_HEADERS))}1"


def generate_engagement_tracker(doc: SowDocument) -> bytes:
    """Return .xlsx bytes for the engagement tracker workbook."""
    wb = Workbook()
    _write_summary_sheet(wb, doc)
    for section in doc.scope_sections:
        _write_section_sheet(wb, section)
    _write_all_deliverables_sheet(wb, doc)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
