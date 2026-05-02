"""Generate a simple issues-tracking Excel workbook for a PCD engagement."""

from __future__ import annotations

import io
import json
import logging

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..templates.excel_styles import (
    HEADER_FILL,
    HEADER_FONT,
    ROW_FILL_EVEN,
    ROW_FILL_ODD,
    SECTION_HEADER_FILL,
    SECTION_HEADER_FONT,
    THIN_BORDER,
    apply_data_row,
    apply_header,
)

log = logging.getLogger(__name__)

_ISSUES_HEADERS = [
    "ID",
    "Summary",
    "Status",
    "Priority",
    "Product",
    "Ownership",
    "Comment",
    "Jira Link",
    "Target Date",
]

_EXTRACT_TOOL: dict = {
    "name": "store_issues",
    "description": "Store extracted issue records.",
    "input_schema": {
        "type": "object",
        "required": ["issues"],
        "properties": {
            "issues": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "summary": {"type": "string"},
                        "status": {
                            "type": "string",
                            "enum": ["Open", "In Progress", "Resolved", "Closed", "Won't Fix"],
                        },
                        "priority": {
                            "type": "string",
                            "enum": ["Critical", "High", "Medium", "Low"],
                        },
                        "product": {"type": "string"},
                        "ownership": {"type": "string"},
                        "comment": {"type": "string"},
                        "jira_link": {"type": "string"},
                    },
                    "required": ["summary"],
                },
            }
        },
    },
}


def _parse_issues_with_claude(raw_text: str) -> list[dict]:
    """Use Claude to extract a structured issues list from raw text."""
    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=4096,
        system=(
            "You are a data extraction assistant. Extract all issues, bugs, or action items "
            "from the provided text and call store_issues with structured records."
        ),
        tools=[_EXTRACT_TOOL],
        tool_choice={"type": "tool", "name": "store_issues"},
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": raw_text,
                        "cache_control": {"type": "ephemeral"},
                    },
                    {"type": "text", "text": "Extract all issues and call store_issues."},
                ],
            }
        ],
    )
    tool_block = next(
        (b for b in response.content if b.type == "tool_use" and b.name == "store_issues"),
        None,
    )
    if tool_block is None:
        log.warning("Claude did not return a store_issues call — returning empty list")
        return []
    return tool_block.input.get("issues", [])


def _parse_issues(data: str) -> list[dict]:
    """Parse issues from JSON string or raw text via Claude."""
    stripped = data.strip()
    if stripped.startswith("[") or stripped.startswith("{"):
        try:
            parsed = json.loads(stripped)
            if isinstance(parsed, list):
                return parsed
            if isinstance(parsed, dict) and "issues" in parsed:
                return parsed["issues"]
        except json.JSONDecodeError:
            log.debug("JSON parse failed — falling back to Claude extraction")
    return _parse_issues_with_claude(data)


def _write_summary_sheet(wb: Workbook, customer_name: str, issues: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"

    # Count by status
    status_counts: dict[str, int] = {}
    priority_counts: dict[str, int] = {}
    for issue in issues:
        status = issue.get("status", "Open")
        priority = issue.get("priority", "Medium")
        status_counts[status] = status_counts.get(status, 0) + 1
        priority_counts[priority] = priority_counts.get(priority, 0) + 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    row = 1
    # Title
    title_cell = ws.cell(row=row, column=1, value=f"Issues Tracker — {customer_name}")
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill = HEADER_FILL
    title_cell.border = THIN_BORDER
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 22
    row += 2

    # Status summary
    ws.cell(row=row, column=1, value="By Status").font = SECTION_HEADER_FONT
    ws.cell(row=row, column=1).fill = SECTION_HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = SECTION_HEADER_FONT
    ws.cell(row=row, column=2).fill = SECTION_HEADER_FILL
    row += 1
    for status, count in sorted(status_counts.items()):
        ws.cell(row=row, column=1, value=status).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        row += 1

    row += 1
    # Priority summary
    ws.cell(row=row, column=1, value="By Priority").font = SECTION_HEADER_FONT
    ws.cell(row=row, column=1).fill = SECTION_HEADER_FILL
    ws.cell(row=row, column=2, value="Count").font = SECTION_HEADER_FONT
    ws.cell(row=row, column=2).fill = SECTION_HEADER_FILL
    row += 1
    for priority, count in sorted(priority_counts.items()):
        ws.cell(row=row, column=1, value=priority).border = THIN_BORDER
        ws.cell(row=row, column=2, value=count).border = THIN_BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Total Issues").border = THIN_BORDER
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(issues)).border = THIN_BORDER


def _write_issues_sheet(wb: Workbook, issues: list[dict]) -> None:
    ws = wb.create_sheet(title="Issue_tracker")
    apply_header(ws, _ISSUES_HEADERS)

    col_widths = {
        "ID": 8, "Summary": 45, "Status": 14, "Priority": 12,
        "Product": 18, "Ownership": 16, "Comment": 40, "Jira Link": 30, "Target Date": 14,
    }
    for col_idx, header in enumerate(_ISSUES_HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 20)

    for idx, issue in enumerate(issues, 1):
        values = [
            f"I-{idx:03d}",
            issue.get("summary", ""),
            issue.get("status", "Open"),
            issue.get("priority", "Medium"),
            issue.get("product", ""),
            issue.get("ownership", "Platform9"),
            issue.get("comment", ""),
            issue.get("jira_link", ""),
            "",  # Target Date
        ]
        apply_data_row(ws, row_idx=idx + 1, values=values, is_odd=(idx % 2 == 1))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_ISSUES_HEADERS))}1"


def _write_bug_sheet(wb: Workbook, issues: list[dict]) -> None:
    """Bug_reported sheet — separate sheet that mirrors all issues (all are potential bugs)."""
    ws = wb.create_sheet(title="Bug_reported")
    bug_headers = ["ID", "Summary", "Priority", "Status", "Product", "Comment", "Jira Link"]
    apply_header(ws, bug_headers)

    col_widths = {
        "ID": 8, "Summary": 45, "Priority": 12, "Status": 14,
        "Product": 18, "Comment": 40, "Jira Link": 30,
    }
    for col_idx, header in enumerate(bug_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 20)

    for idx, issue in enumerate(issues, 1):
        values = [
            f"I-{idx:03d}",
            issue.get("summary", ""),
            issue.get("priority", "Medium"),
            issue.get("status", "Open"),
            issue.get("product", ""),
            issue.get("comment", ""),
            issue.get("jira_link", ""),
        ]
        apply_data_row(ws, row_idx=idx + 1, values=values, is_odd=(idx % 2 == 1))

    ws.freeze_panes = "A2"


def _write_open_actions_sheet(wb: Workbook, issues: list[dict]) -> None:
    """Open Actions sheet — filtered view of issues with non-closed status."""
    ws = wb.create_sheet(title="Open Actions")
    action_headers = ["ID", "Summary", "Status", "Priority", "Ownership", "Target Date", "Comment"]
    apply_header(ws, action_headers)

    col_widths = {
        "ID": 8, "Summary": 45, "Status": 14, "Priority": 12,
        "Ownership": 16, "Target Date": 14, "Comment": 40,
    }
    for col_idx, header in enumerate(action_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 20)

    open_statuses = {"Open", "In Progress"}
    data_row = 2
    for idx, issue in enumerate(issues, 1):
        status = issue.get("status", "Open")
        if status not in open_statuses:
            continue
        values = [
            f"I-{idx:03d}",
            issue.get("summary", ""),
            status,
            issue.get("priority", "Medium"),
            issue.get("ownership", "Platform9"),
            "",  # Target Date
            issue.get("comment", ""),
        ]
        apply_data_row(ws, row_idx=data_row, values=values, is_odd=(data_row % 2 == 0))
        data_row += 1

    ws.freeze_panes = "A2"


def generate_issues_tracker_from_text(customer_name: str, data: str) -> bytes:
    """
    Parse *data* (JSON or raw text) into issues and build a 4-sheet Excel workbook.

    Sheets: Summary, Issue_tracker, Bug_reported, Open Actions.

    Returns:
        Raw .xlsx bytes.
    """
    log.info("Parsing issues for customer: %s", customer_name)
    issues = _parse_issues(data)
    log.info("Parsed %d issues", len(issues))

    wb = Workbook()
    _write_summary_sheet(wb, customer_name, issues)
    _write_issues_sheet(wb, issues)
    _write_bug_sheet(wb, issues)
    _write_open_actions_sheet(wb, issues)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
