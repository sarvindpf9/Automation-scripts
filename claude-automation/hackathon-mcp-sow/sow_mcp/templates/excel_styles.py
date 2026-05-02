from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Platform9 brand colours
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ROW_FILL_ODD = PatternFill("solid", fgColor="DEEAF1")
ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")

SECTION_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")

_THIN = Side(style="thin")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Default column widths by header name; unknown headers fall back to 20.
COLUMN_WIDTHS: dict[str, int] = {
    "ID": 12,
    "Section": 10,
    "Section Title": 38,
    "Title": 45,
    "Type": 16,
    "Owner": 14,
    "Est. Hours": 12,
    "Target Date": 14,
    "Status": 14,
    "Notes": 40,
    "Label": 22,
    "Value": 60,
}


def apply_header(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 20)
    ws.row_dimensions[row].height = 28


def apply_data_row(ws: Worksheet, row_idx: int, values: list, is_odd: bool) -> None:
    fill = ROW_FILL_ODD if is_odd else ROW_FILL_EVEN
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
