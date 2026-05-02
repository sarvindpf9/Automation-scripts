#!/usr/bin/env python3
"""
Manual test script for the Platform9 PCD MCP server.

Two modes:

  --direct   Call the Python generator functions directly (fast, no subprocess).
             Tests tracker and docs_reader without an API key.

  --mcp      Spawn the real MCP server as a subprocess and communicate over
             stdin/stdout using JSON-RPC 2.0 (the actual MCP wire protocol).
             Tests tools/list, then calls generate_tracker.

  --all      Run both modes, then attempt LLM-backed tools if ANTHROPIC_API_KEY is set.

Usage:
  python test_manual.py --direct
  python test_manual.py --mcp
  ANTHROPIC_API_KEY=sk-ant-... python test_manual.py --all
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).parent
VENV_PYTHON = ROOT / ".venv" / "bin" / "python3.14"
PYTHON = str(VENV_PYTHON) if VENV_PYTHON.exists() else sys.executable

# ── Sample data ───────────────────────────────────────────────────────────────

SAMPLE_ISSUES_JSON = json.dumps([
    {
        "summary": "SSO with ADFS integration pending",
        "status": "In-discussion",
        "priority": "high",
        "product": "PCD-V",
        "ownership": "Platform9",
        "comment": "ECS-171 raised for engineering assistance",
        "jira_link": "https://platform9.atlassian.net/browse/ECS-171",
    },
    {
        "summary": "Commvault backup integration not started",
        "status": "Not_started",
        "priority": "high",
        "product": "PCD-V",
        "ownership": "Platform9",
        "comment": "",
        "jira_link": None,
    },
    {
        "summary": "VMHA waiting state after host onboarding",
        "status": "Done",
        "priority": "high",
        "product": "PCD-V",
        "ownership": "Platform9",
        "comment": "Temp workaround applied; bug PCD-5795 filed",
        "jira_link": "https://platform9.atlassian.net/browse/PCD-5795",
    },
])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ok(label: str) -> None:
    print(f"  \033[32m✓\033[0m  {label}")


def _fail(label: str, detail: str = "") -> None:
    print(f"  \033[31m✗\033[0m  {label}")
    if detail:
        print(f"       {detail}")


def _section(title: str) -> None:
    print(f"\n\033[1m{title}\033[0m")
    print("─" * 60)


# ── Direct mode ───────────────────────────────────────────────────────────────

def test_direct() -> None:
    _section("Direct mode — calling Python functions without MCP overhead")

    # 1. docs_reader
    try:
        sys.path.insert(0, str(ROOT))
        from sow_mcp.docs_reader import read_docs_dir
        docs = read_docs_dir(str(ROOT / "docs"))
        assert len(docs) > 0, "no docs found"
        total_chars = sum(len(d["text"]) for d in docs)
        _ok(f"docs_reader: read {len(docs)} files ({total_chars:,} chars total)")
        for d in docs:
            print(f"       {d['filename']:55s}  {len(d['text']):>7,} chars")
    except Exception as exc:
        _fail("docs_reader", str(exc))

    # 2. issues tracker (no API key)
    try:
        import io
        from openpyxl import load_workbook
        from sow_mcp.generators.issues_tracker import generate_issues_tracker_from_text
        from sow_mcp.output_writer import OutputWriter

        xlsx = generate_issues_tracker_from_text("TestCustomer", SAMPLE_ISSUES_JSON)
        writer = OutputWriter("TestCustomer")
        path = writer.write_bytes("TestCustomer_issues_tracker.xlsx", xlsx)

        wb = load_workbook(io.BytesIO(xlsx))
        _ok(f"issues_tracker: {len(wb.sheetnames)} sheets → {path}")
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = ws.max_row - 1  # subtract header
            print(f"       Sheet '{sheet}': {max(rows, 0)} data rows")
    except Exception as exc:
        _fail("issues_tracker", str(exc))
        raise

    # 3. SOW engagement tracker (no API key)
    try:
        import io
        import dacite
        from openpyxl import load_workbook
        from sow_mcp.models import (
            Datacenter, Deliverable, ScopeSection, SowDocument,
        )
        from sow_mcp.generators.engagement_tracker import generate_engagement_tracker
        from sow_mcp.output_writer import OutputWriter

        doc = SowDocument(
            customer="TestCustomer",
            project_title="PCD Platinum Deployment",
            prepared_by="Platform9 Strategic Customer Engineering",
            datacenters=[
                Datacenter(name="Canary-DC", role="canary", hypervisor_count=4),
                Datacenter(name="Prod-DC", role="production", hypervisor_count=50),
            ],
            phases=["Canary", "Production", "MaaS", "Migration"],
            integrations=["NetApp AFF C80", "Microsoft AD"],
            vm_count=500,
            hypervisor_count=50,
            migration_tool="vJailbreak",
            scope_sections=[
                ScopeSection(
                    id="3.1",
                    title="PCD Canary Deployment",
                    activities=["Deploy PCD SaaS", "Validate storage"],
                    deliverables=[
                        Deliverable(
                            id="D-3.1-1", section_id="3.1",
                            title="Canary Build MOP",
                            type="mop", owner="Platform9", estimated_hours=16.0,
                        ),
                    ],
                ),
            ],
        )
        xlsx = generate_engagement_tracker(doc)
        writer = OutputWriter("TestCustomer")
        path = writer.write_bytes("TestCustomer_engagement_tracker.xlsx", xlsx)
        wb = load_workbook(io.BytesIO(xlsx))
        _ok(f"engagement_tracker: {len(wb.sheetnames)} sheets → {path}")
    except Exception as exc:
        _fail("engagement_tracker", str(exc))

    # 4. LLM-backed tools (only if API key present)
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("\n  (skipping LLM tools — ANTHROPIC_API_KEY not set)")
        return

    # 4a. answer_query
    try:
        from sow_mcp.generators.query_answerer import answer_query
        answer = answer_query("What are the prerequisites for self-hosted Platform9 PCD installation?")
        assert len(answer) > 100
        _ok(f"answer_query: got {len(answer)} char answer")
        print(f"\n  Preview:\n  {answer[:300].replace(chr(10), chr(10)+'  ')}\n  ...")
    except Exception as exc:
        _fail("answer_query", str(exc))

    # 4b. build_automation
    try:
        from sow_mcp.generators.automation_builder import build_automation
        code = build_automation("list all Nova instances across all projects", language="python")
        assert "import" in code or "def " in code
        _ok(f"build_automation: got {len(code)} char script")
    except Exception as exc:
        _fail("build_automation", str(exc))

    # 4c. procedure_doc (full stack)
    try:
        from sow_mcp.docs_reader import read_docs_dir
        from sow_mcp.generators.procedure_doc import generate_procedure_doc
        from sow_mcp.output_writer import OutputWriter

        docs = read_docs_dir(str(ROOT / "docs"))
        writer = OutputWriter("TestCustomer")
        md_path, pdf_path = generate_procedure_doc("TestCustomer", docs, writer)
        assert md_path.exists() and pdf_path.exists()
        content = md_path.read_text()
        for day in ("Day 0", "Day 1", "Day 2"):
            assert day in content, f"Missing {day} in output"
        _ok(f"procedure_doc: {md_path.name} + {pdf_path.name}")
        print(f"       Markdown: {md_path}")
        print(f"       PDF:      {pdf_path}")
    except Exception as exc:
        _fail("procedure_doc", str(exc))


# ── MCP protocol mode ─────────────────────────────────────────────────────────

def _send(proc: subprocess.Popen, msg: dict) -> None:
    line = json.dumps(msg) + "\n"
    proc.stdin.write(line.encode())
    proc.stdin.flush()


def _recv(proc: subprocess.Popen, timeout: float = 10.0) -> dict | None:
    """Read one JSON-RPC line from the server stdout."""
    start = time.monotonic()
    while time.monotonic() - start < timeout:
        line = proc.stdout.readline()
        if not line:
            time.sleep(0.05)
            continue
        line = line.strip()
        if not line:
            continue
        try:
            return json.loads(line)
        except json.JSONDecodeError:
            continue  # server may emit log lines to stdout on some versions
    return None


def test_mcp() -> None:
    _section("MCP protocol mode — JSON-RPC 2.0 over stdio (real wire format)")

    env = {**os.environ, "PYTHONPATH": str(ROOT)}
    proc = subprocess.Popen(
        [PYTHON, "-m", "sow_mcp.server"],
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        cwd=str(ROOT),
        env=env,
    )

    try:
        # ── 1. initialize ─────────────────────────────────────────────────────
        _send(proc, {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {"name": "manual-test", "version": "0.1"},
            },
        })
        resp = _recv(proc)
        if resp and resp.get("result", {}).get("protocolVersion"):
            _ok(f"initialize: protocol {resp['result']['protocolVersion']}, "
                f"server '{resp['result']['serverInfo']['name']}'")
        else:
            _fail("initialize", repr(resp))
            return

        # ── 2. initialized notification (required before any tool call) ───────
        _send(proc, {"jsonrpc": "2.0", "method": "notifications/initialized"})

        # ── 3. tools/list ─────────────────────────────────────────────────────
        _send(proc, {"jsonrpc": "2.0", "id": 2, "method": "tools/list"})
        resp = _recv(proc)
        if resp and "result" in resp:
            tools = [t["name"] for t in resp["result"].get("tools", [])]
            _ok(f"tools/list: {tools}")
        else:
            _fail("tools/list", repr(resp))
            return

        # ── 4. tools/call — generate_tracker (no API key needed) ─────────────
        _send(proc, {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "generate_tracker",
                "arguments": {
                    "customer_name": "MCPTest",
                    "data": SAMPLE_ISSUES_JSON,
                    "tracker_type": "issues",
                },
            },
        })
        resp = _recv(proc, timeout=30.0)
        if resp and "result" in resp:
            content = resp["result"].get("content", [{}])
            text = content[0].get("text", "") if content else ""
            if ".xlsx" in text:
                _ok(f"tools/call generate_tracker: {text.strip()}")
            else:
                _fail("tools/call generate_tracker: unexpected response", text[:200])
        else:
            err = resp.get("error") if resp else "timeout"
            _fail("tools/call generate_tracker", repr(err))

        # ── 5. tools/call — answer_query (needs API key) ──────────────────────
        if os.getenv("ANTHROPIC_API_KEY"):
            _send(proc, {
                "jsonrpc": "2.0",
                "id": 4,
                "method": "tools/call",
                "params": {
                    "name": "answer_query",
                    "arguments": {
                        "question": "What ports does Platform9 PCD require open for the management plane?",
                    },
                },
            })
            resp = _recv(proc, timeout=60.0)
            if resp and "result" in resp:
                content = resp["result"].get("content", [{}])
                text = content[0].get("text", "") if content else ""
                _ok(f"tools/call answer_query: got {len(text)} char answer")
            else:
                err = resp.get("error") if resp else "timeout"
                _fail("tools/call answer_query", repr(err))
        else:
            print("  (skipping answer_query — ANTHROPIC_API_KEY not set)")

    finally:
        proc.stdin.close()
        proc.wait(timeout=5)


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--direct", action="store_true", help="Call Python functions directly")
    group.add_argument("--mcp",    action="store_true", help="Test via MCP JSON-RPC protocol over stdio")
    group.add_argument("--all",    action="store_true", help="Run both modes")
    args = parser.parse_args()

    api_key = os.getenv("ANTHROPIC_API_KEY")
    print(f"ANTHROPIC_API_KEY : {'set ✓' if api_key else 'not set — LLM tools will be skipped'}")
    print(f"Python            : {PYTHON}")
    print(f"Working directory : {ROOT}")

    if args.direct or args.all:
        test_direct()
    if args.mcp or args.all:
        test_mcp()

    print()


if __name__ == "__main__":
    main()
