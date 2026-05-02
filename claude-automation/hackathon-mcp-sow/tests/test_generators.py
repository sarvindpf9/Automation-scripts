"""
Generator tests.
- Engagement tracker (SOW-based): pure openpyxl — always runs, no API key needed.
- Issues tracker structure: openpyxl — always runs (uses JSON path, no LLM).
- docs_reader: pdfplumber — runs when fixture is present.
- procedure_doc / query_answerer / automation_builder: require ANTHROPIC_API_KEY.
"""
import io
import json
import os
import pytest
from pathlib import Path

from openpyxl import load_workbook

from sow_mcp.models import Datacenter, Deliverable, ScopeSection, SowDocument
from sow_mcp.generators.engagement_tracker import generate_engagement_tracker
from sow_mcp.docs_reader import read_docs_dir, read_file

FIXTURE_PDF = Path(__file__).parent / "fixtures" / "NTT-SOW.pdf"
DOCS_DIR = Path(__file__).parent.parent / "docs"

_has_fixture = FIXTURE_PDF.exists()
_has_docs = DOCS_DIR.exists() and any(DOCS_DIR.glob("*.pdf"))
_has_api_key = bool(os.getenv("ANTHROPIC_API_KEY"))

needs_fixture = pytest.mark.skipif(not _has_fixture, reason="NTT-SOW.pdf fixture not present")
needs_docs = pytest.mark.skipif(not _has_docs, reason="docs/ directory has no PDFs")
needs_llm = pytest.mark.skipif(
    not _has_api_key,
    reason="ANTHROPIC_API_KEY not set",
)


# ── Shared SOW fixture ────────────────────────────────────────────────────────

@pytest.fixture()
def sample_doc() -> SowDocument:
    return SowDocument(
        customer="Acme Corp",
        project_title="PCD Platinum Deployment",
        prepared_by="Platform9 Strategic Customer Engineering",
        datacenters=[
            Datacenter(name="London", role="canary", hypervisor_count=4),
            Datacenter(name="Frankfurt", role="production", hypervisor_count=50),
        ],
        phases=["Canary", "Production", "MaaS", "Migration", "Training"],
        integrations=["NetApp AFF C80", "Microsoft AD", "Nagios"],
        vm_count=500,
        hypervisor_count=50,
        migration_tool="vJailbreak",
        scope_sections=[
            ScopeSection(
                id="3.1",
                title="Assessment, Design & Build PCD — Canary Environment",
                activities=[
                    "Deploy PCD with SaaS control plane in London DC",
                    "Validate network topology and storage integration",
                ],
                deliverables=[
                    Deliverable(
                        id="D-3.1-1", section_id="3.1",
                        title="Canary Environment Build MOP",
                        type="mop", owner="Platform9", estimated_hours=16.0,
                    ),
                    Deliverable(
                        id="D-3.1-2", section_id="3.1",
                        title="Canary Validation Report",
                        type="report", owner="Platform9", estimated_hours=4.0,
                    ),
                ],
            ),
            ScopeSection(
                id="3.2",
                title="Key Integrations",
                activities=["Integrate NetApp AFF C80 via fibre-channel"],
                deliverables=[],
            ),
        ],
        assumptions=["Customer provides rack access", "Network VLANs pre-provisioned"],
        out_of_scope=["Application-level testing", "Licensing procurement"],
    )


# ── SOW engagement tracker (no LLM) ──────────────────────────────────────────

def test_tracker_returns_bytes(sample_doc):
    assert len(generate_engagement_tracker(sample_doc)) > 1000


def test_tracker_is_valid_xlsx(sample_doc):
    assert load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc))) is not None


def test_tracker_has_summary_sheet(sample_doc):
    wb = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))
    assert "Summary" in wb.sheetnames


def test_tracker_has_section_sheets(sample_doc):
    sheets = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc))).sheetnames
    assert any("3.1" in s for s in sheets)
    assert any("3.2" in s for s in sheets)


def test_tracker_has_all_deliverables_sheet(sample_doc):
    wb = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))
    assert "All Deliverables" in wb.sheetnames


def test_tracker_section_sheet_has_deliverables(sample_doc):
    wb = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))
    ws = wb[next(s for s in wb.sheetnames if "3.1" in s)]
    assert ws.cell(row=2, column=1).value == "D-3.1-1"


def test_tracker_empty_section_writes_placeholder(sample_doc):
    wb = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))
    ws = wb[next(s for s in wb.sheetnames if "3.2" in s)]
    assert "[REVIEW]" in (ws.cell(row=2, column=2).value or "")


def test_tracker_summary_contains_customer(sample_doc):
    wb = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))
    values = [wb["Summary"].cell(row=r, column=2).value for r in range(1, 20)]
    assert "Acme Corp" in values


def test_tracker_all_deliverables_count(sample_doc):
    ws = load_workbook(io.BytesIO(generate_engagement_tracker(sample_doc)))["All Deliverables"]
    assert ws.cell(row=2, column=1).value is not None
    assert ws.cell(row=3, column=1).value is not None
    assert ws.cell(row=4, column=1).value is None


# ── docs_reader (no LLM) ─────────────────────────────────────────────────────

@needs_fixture
def test_read_file_pdf_returns_text():
    text = read_file(str(FIXTURE_PDF))
    assert len(text) > 200
    assert "Platform9" in text


@needs_docs
def test_read_docs_dir_returns_list():
    docs = read_docs_dir(str(DOCS_DIR))
    assert isinstance(docs, list)
    assert len(docs) > 0


@needs_docs
def test_read_docs_dir_each_has_text():
    for doc in read_docs_dir(str(DOCS_DIR)):
        assert doc["filename"]
        assert len(doc["text"]) > 50, f"{doc['filename']} returned very short text"


def test_read_file_missing_raises():
    with pytest.raises(FileNotFoundError):
        read_file("/nonexistent/file.pdf")


def test_read_file_unsupported_type_raises():
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as f:
        f.write(b"hello"); name = f.name
    try:
        with pytest.raises(ValueError, match="Unsupported"):
            read_file(name)
    finally:
        os.unlink(name)


# ── issues_tracker structure (no LLM — JSON path) ────────────────────────────

def test_issues_tracker_from_json():
    from sow_mcp.generators.issues_tracker import generate_issues_tracker_from_text
    issues_json = json.dumps([
        {
            "summary": "SSO with ADFS not configured",
            "status": "In-discussion",
            "priority": "high",
            "product": "PCD-V",
            "ownership": "Platform9",
            "comment": "ECS-171 raised",
            "jira_link": "https://platform9.atlassian.net/browse/ECS-171",
        },
        {
            "summary": "Commvault integration pending",
            "status": "Not_started",
            "priority": "high",
            "product": "PCD-V",
            "ownership": "Platform9",
            "comment": "",
            "jira_link": None,
        },
    ])
    result = generate_issues_tracker_from_text("TestCo", issues_json)
    assert isinstance(result, bytes)
    wb = load_workbook(io.BytesIO(result))
    assert "Issues" in wb.sheetnames or any("Issue" in s for s in wb.sheetnames)


# ── LLM-backed tests ──────────────────────────────────────────────────────────

@needs_llm
@needs_docs
def test_procedure_doc_returns_markdown():
    from sow_mcp.docs_reader import read_docs_dir
    from sow_mcp.generators.procedure_doc import generate_procedure_doc
    from sow_mcp.output_writer import OutputWriter
    docs = read_docs_dir(str(DOCS_DIR))
    writer = OutputWriter("test_acme", base_dir=Path("/tmp/sow_mcp_test"))
    md_path, pdf_path = generate_procedure_doc("Acme Corp", docs, writer)
    assert md_path.exists()
    assert pdf_path.exists()
    content = md_path.read_text()
    assert "Day 0" in content
    assert "Day 1" in content
    assert "Day 2" in content


@needs_llm
def test_answer_query_returns_text():
    from sow_mcp.generators.query_answerer import answer_query
    result = answer_query("What are the prerequisites for installing Platform9 PCD?")
    assert isinstance(result, str)
    assert len(result) > 100


@needs_llm
def test_build_automation_returns_code():
    from sow_mcp.generators.automation_builder import build_automation
    result = build_automation("list all Nova instances in all projects", language="python")
    assert "import" in result or "def " in result
    assert "OS_AUTH_URL" in result or "auth_url" in result
